
import os

import openpyxl

from handlers.secondary_functions import make_file
from CONSTANTS import *


def parse_settings(data):
    make_file(data, FILES_PATH + "settings.xlsx")

    settings_book = openpyxl.load_workbook(FILES_PATH + "settings.xlsx")
    settings_sheet = settings_book['Настройки']
    names_values = [v[0].value for v in settings_sheet['A2':'A12']]
    params_values = [v[0].value for v in settings_sheet['B2':'B12']]

    os.remove(FILES_PATH + "settings.xlsx")

    params = {key: value for key, value in list(zip(names_values, params_values))}

    # params['pattern'] = data["Шаблон"]
    # params['org_name'] = data['Наименование организации']
    # params['inn'] = data['ИНН']
    # params['kpp'] = data['КПП']
    # params['bik'] = data['БИК']
    # params['corr_account'] = data['Корреспондентский счет']
    # params['bank_name'] = data['Наименование банка']
    # params['payment_account'] = data['Расчетный счет']
    # params['service_code'] = data['Код услуги']
    # params['dsk'] = data['Дополнительные параметры ДШК']

    return params


def parse_payers(data):
    """
    This function parses payers' data from table(Name, number, etc.)
    :return: parsed list
    """
    make_file(data, FILES_PATH + "payers.xlsx")

    receipt_book = openpyxl.load_workbook(FILES_PATH + "payers.xlsx")

    sheet = receipt_book['Реестр начислений']
    
    parsed = []

    for row in sheet.iter_rows(min_row=7):
        _tuple = ()
        if all(i.value is None for i in row):
            continue
        for cell in row:
            
            if '.' in str(cell.value) and len(str(cell.value).split('.')[-1]) == 1:
                _tuple += (str(cell.value).strip() + '0',)
            else:
                _tuple += (str(cell.value).strip(),)
        parsed.append(_tuple)
    print(parsed)

    # os.remove("payers.xlsx")

    return parsed


def parse_emails(data):
    """
    This function parses payers' emails from table.
    :return: parsed list
    """
    make_file(data, FILES_PATH + "emails.xlsx")

    receipt_book = openpyxl.load_workbook(FILES_PATH + "emails.xlsx")

    sheet = receipt_book['emails']

    parsed = {row[0].value.strip(): row[1].value.strip()
              for row in sheet.iter_rows(min_row=2) if None not in (row[0].value, row[1].value)}

    # os.remove("emails.xlsx")

    return parsed
